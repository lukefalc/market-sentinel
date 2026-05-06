"""Generate the market-sentinel Excel report.

Run this script from the project root with:

    python3 scripts/generate_excel_report.py
"""

import sys
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = PROJECT_ROOT / "src"

if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from market_sentinel.database.connection import open_duckdb_connection  # noqa: E402
from market_sentinel.database.schema import initialise_database_schema  # noqa: E402
from market_sentinel.reports.excel_report import (  # noqa: E402
    EXPECTED_WORKSHEET_TITLES,
    generate_excel_report,
)
from market_sentinel.utils.timing import timed_step  # noqa: E402


def main() -> None:
    """Generate an Excel workbook from the local DuckDB database."""
    connection = None

    try:
        with timed_step("Generate Excel report"):
            connection = open_duckdb_connection()
            initialise_database_schema(connection)
            output_path = generate_excel_report(connection)
    except (RuntimeError, ValueError, FileNotFoundError) as error:
        print(f"Excel report generation failed: {error}", file=sys.stderr)
        raise SystemExit(1) from error
    finally:
        if connection is not None:
            connection.close()

    print(f"Excel report saved: {output_path}")
    print("Worksheet tabs:")
    for sheet_name in _read_saved_worksheet_titles(output_path):
        print(f"- {sheet_name}")


def _read_saved_worksheet_titles(output_path: Path) -> list:
    """Read worksheet titles from the saved Excel workbook."""
    workbook = load_workbook(output_path, read_only=True)
    sheet_names = workbook.sheetnames

    if sheet_names != EXPECTED_WORKSHEET_TITLES:
        raise RuntimeError(
            "The saved Excel report does not have the expected worksheet tabs. "
            f"Expected: {EXPECTED_WORKSHEET_TITLES}. Found: {sheet_names}."
        )

    return sheet_names


if __name__ == "__main__":
    main()
