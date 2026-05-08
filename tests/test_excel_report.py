"""Tests for Excel report generation."""

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts import generate_excel_report as generate_excel_report_script
from market_sentinel.database.connection import open_duckdb_connection
from market_sentinel.database.schema import initialise_database_schema
from market_sentinel.reports.excel_report import (
    EXPECTED_WORKSHEET_TITLES,
    REVIEW_DECISION_VALUES,
    TRADE_CANDIDATE_POSITION_HEADERS,
    TRADE_CANDIDATE_REVIEW_HEADERS,
    _position_sizing_values,
    apply_trade_candidate_position_sizing,
    generate_excel_report,
)


def write_settings(config_dir: Path, database_path: Path) -> None:
    """Create a minimal settings file pointing at a test database."""
    config_dir.mkdir(parents=True, exist_ok=True)
    config_dir.joinpath("settings.yaml").write_text(
        f"database_path: {database_path}\n",
        encoding="utf-8",
    )


def write_report_settings(
    config_dir: Path,
    database_path: Path,
    excel_dir: str,
    max_rows_per_sheet: int = 50000,
    recent_days: int = 10,
) -> None:
    """Create settings with a custom Excel report output folder."""
    config_dir.mkdir(parents=True, exist_ok=True)
    config_dir.joinpath("settings.yaml").write_text(
        "\n".join(
            [
                f"database_path: {database_path}",
                f"excel_max_rows_per_sheet: {max_rows_per_sheet}",
                f"excel_moving_average_recent_days: {recent_days}",
                "report_outputs:",
                f"  excel_dir: {excel_dir}",
            ]
        ),
        encoding="utf-8",
    )


def open_test_database(tmp_path: Path):
    """Open a temporary DuckDB database with the project schema."""
    config_dir = tmp_path / "config"
    database_path = tmp_path / "data" / "market_sentinel.duckdb"
    write_settings(config_dir, database_path)
    connection = open_duckdb_connection(config_dir)
    initialise_database_schema(connection)
    return connection


def insert_report_data(connection) -> None:
    """Insert fake database rows for the Excel report."""
    connection.execute(
        """
        INSERT INTO securities (
            security_id,
            ticker,
            name,
            market,
            region,
            currency,
            sector
        )
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        [1, "AAA", "Example A", "S&P 500", "United States", "USD", "Technology"],
    )
    connection.execute(
        """
        INSERT INTO securities (
            security_id,
            ticker,
            name,
            market,
            region,
            currency,
            sector
        )
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        [2, "BBB", "Example B", "FTSE 350", "United Kingdom", "GBP", "Energy"],
    )
    connection.execute(
        """
        INSERT INTO daily_prices (
            price_id,
            security_id,
            price_date,
            open_price,
            high_price,
            low_price,
            close_price,
            adjusted_close_price,
            volume
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [1, 1, "2026-05-01", 10.0, 12.0, 9.0, 11.0, 10.8, 1000],
    )
    connection.execute(
        """
        INSERT INTO moving_average_signals (
            signal_id,
            security_id,
            signal_date,
            moving_average_period_days,
            moving_average_value,
            signal_type
        )
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        [1, 1, "2026-05-01", 7, 10.5, "SMA"],
    )
    connection.execute(
        """
        INSERT INTO moving_average_signals (
            signal_id,
            security_id,
            signal_date,
            moving_average_period_days,
            moving_average_value,
            comparison_period_days,
            comparison_moving_average_value,
            signal_type,
            crossover_direction
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            2,
            1,
            "2026-05-01",
            7,
            10.5,
            30,
            10.0,
            "BULLISH_CROSSOVER",
            "BULLISH_CROSSOVER",
        ],
    )
    connection.execute(
        """
        INSERT INTO moving_average_signals (
            signal_id,
            security_id,
            signal_date,
            moving_average_period_days,
            moving_average_value,
            comparison_period_days,
            comparison_moving_average_value,
            signal_type,
            crossover_direction
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            4,
            2,
            "2026-04-20",
            7,
            9.0,
            30,
            10.0,
            "BEARISH_CROSSOVER",
            "BEARISH_CROSSOVER",
        ],
    )
    connection.execute(
        """
        INSERT INTO moving_average_signals (
            signal_id,
            security_id,
            signal_date,
            moving_average_period_days,
            moving_average_value,
            comparison_period_days,
            comparison_moving_average_value,
            signal_type,
            crossover_direction
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            3,
            2,
            "2026-05-01",
            7,
            9.5,
            30,
            10.0,
            "BEARISH_CROSSOVER",
            "BEARISH_CROSSOVER",
        ],
    )
    connection.execute(
        """
        INSERT INTO dividend_metrics (
            metric_id,
            security_id,
            metric_date,
            trailing_annual_dividend,
            dividend_yield,
            annual_dividend_cash_per_10000,
            dividend_risk_flag,
            dividend_risk_reason
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            1,
            1,
            "2026-05-01",
            2.5,
            0.08,
            800.0,
            "DIVIDEND_TRAP_RISK",
            "Dividend yield is above 7%.",
        ],
    )
    connection.execute(
        """
        INSERT INTO dividend_metrics (
            metric_id,
            security_id,
            metric_date,
            trailing_annual_dividend,
            dividend_yield,
            annual_dividend_cash_per_10000
        )
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        [2, 2, "2026-05-01", 1.0, 0.03, 300.0],
    )


def test_generate_excel_report_creates_expected_workbook(tmp_path: Path) -> None:
    """Excel report generation should create the expected workbook tabs."""
    connection = open_test_database(tmp_path)
    output_dir = tmp_path / "outputs" / "excel"

    try:
        insert_report_data(connection)
        output_path = generate_excel_report(
            connection,
            output_dir=output_dir,
            report_date=date(2026, 5, 3),
        )
    finally:
        connection.close()

    workbook = load_workbook(output_path)

    assert output_path == output_dir / "market_sentinel_report_2026-05-03.xlsx"
    assert output_path.exists()
    assert workbook.sheetnames == EXPECTED_WORKSHEET_TITLES
    summary_values = {
        workbook["Summary"][f"A{row}"].value: workbook["Summary"][f"B{row}"].value
        for row in range(1, workbook["Summary"].max_row + 1)
    }

    assert summary_values["Securities"] == 2
    assert summary_values["Securities With Dividend Metrics"] == 2
    assert summary_values["Dividend Risk Flags"] == 1
    assert workbook["Securities"]["A2"].value == "AAA"
    assert workbook["Latest Prices"]["A2"].value == "AAA"
    assert workbook["Latest Prices"]["B1"].value == "Market"
    assert workbook["Latest Prices"]["B2"].value == "S&P 500"
    assert workbook["Moving Averages"]["A2"].value == "AAA"
    assert workbook["Moving Averages"]["B1"].value == "Market"
    assert workbook["Recent Moving Averages"]["A2"].value == "AAA"
    assert workbook["Recent Moving Averages"]["B1"].value == "Market"
    assert workbook["Recent Crossovers"]["A2"].value == "AAA"
    assert workbook["Recent Crossovers"]["C1"].value == "Market"
    assert workbook["Recent Crossovers"]["D2"].value == "Bullish"
    assert workbook["Recent Crossovers"]["A3"].value == "BBB"
    assert workbook["Recent Crossovers"].max_row == 3
    assert workbook["Crossover Signals"]["D2"].value == "Bullish"
    assert (
        workbook["Crossover Signals"]["E2"].value
        == "7-day trend line crossed above 30-day trend line"
    )
    assert workbook["Crossover Signals"]["F2"].value.date() == date(2026, 5, 1)
    assert workbook["Crossover Signals"]["G2"].value == "2 days ago"
    assert workbook["Crossover Signals"]["A3"].value == "BBB"
    assert workbook["Crossover Signals"]["D3"].value == "Bearish"
    assert (
        workbook["Crossover Signals"]["E3"].value
        == "7-day trend line crossed below 30-day trend line"
    )
    assert workbook["Crossover Signals"]["G4"].value == "13 days ago"
    assert workbook["Dividend Metrics"]["A2"].value == "AAA"
    assert workbook["Dividend Metrics"]["B1"].value == "Market"
    assert workbook["High Dividend Stocks"]["A2"].value == "AAA"
    assert workbook["High Dividend Stocks"]["B1"].value == "Market"
    assert workbook["Dividend Risk Flags"]["B1"].value == "Market"
    assert workbook["Dividend Risk Flags"]["F2"].value == "Dividend yield is above 7%."
    assert workbook["Trade Candidates"]["A1"].value == "Ticker"
    assert workbook["Trade Candidates"]["C1"].value == "Market"
    assert workbook["Trade Candidates"]["E1"].value == "Action Grade"
    assert workbook["Trade Candidates"]["A2"].value == "AAA"
    assert workbook["Trade Candidates"]["E2"].value is not None
    assert workbook["Trade Candidates"].auto_filter.ref is not None
    assert workbook["Trade Candidates"].freeze_panes == "A2"
    trade_candidate_headers = [
        cell.value for cell in workbook["Trade Candidates"][1]
    ]
    assert trade_candidate_headers[-len(TRADE_CANDIDATE_REVIEW_HEADERS) :] == (
        TRADE_CANDIDATE_REVIEW_HEADERS
    )
    assert "Review decision" in trade_candidate_headers
    assert "Planned stop" in trade_candidate_headers
    assert "Position size" in trade_candidate_headers
    for header in TRADE_CANDIDATE_POSITION_HEADERS:
        assert header in trade_candidate_headers
    portfolio_status_column = trade_candidate_headers.index("Portfolio Status") + 1
    planned_entry_column = trade_candidate_headers.index("Planned entry") + 1
    planned_stop_column = trade_candidate_headers.index("Planned stop") + 1
    planned_risk_column = trade_candidate_headers.index("Planned risk %") + 1
    position_size_column = trade_candidate_headers.index("Position size") + 1
    risk_per_unit_column = trade_candidate_headers.index("Risk per unit") + 1
    max_risk_column = trade_candidate_headers.index("Max £ risk") + 1
    position_value_column = trade_candidate_headers.index("Position value") + 1
    review_decision_column = trade_candidate_headers.index("Review decision") + 1
    review_notes_column = trade_candidate_headers.index("Review notes") + 1
    reviewed_date_column = trade_candidate_headers.index("Reviewed date") + 1
    assert workbook["Trade Candidates"].cell(2, portfolio_status_column).value == "New"
    planned_entry = workbook["Trade Candidates"].cell(2, planned_entry_column).value
    planned_stop = workbook["Trade Candidates"].cell(2, planned_stop_column).value
    assert planned_entry
    assert planned_entry.startswith("=")
    assert planned_stop
    assert planned_stop.startswith("=")
    assert workbook["Trade Candidates"].cell(2, planned_risk_column).value == 0.01
    assert (
        str(workbook["Trade Candidates"].cell(2, position_size_column).value)
        .startswith("=IF(")
    )
    assert workbook["Trade Candidates"].cell(2, risk_per_unit_column).value.startswith("=")
    assert workbook["Trade Candidates"].cell(2, max_risk_column).value.startswith("=")
    assert workbook["Trade Candidates"].cell(2, position_value_column).value.startswith("=")
    assert workbook["Trade Candidates"].cell(2, review_decision_column).value is None
    assert workbook["Trade Candidates"].cell(2, review_notes_column).value is None
    assert workbook["Trade Candidates"].cell(2, reviewed_date_column).value is None


def test_trade_candidates_sheet_has_review_decision_values(
    tmp_path: Path,
) -> None:
    """Trade Candidates should offer suggested daily review decisions."""
    connection = open_test_database(tmp_path)
    output_dir = tmp_path / "outputs" / "excel"

    try:
        insert_report_data(connection)
        output_path = generate_excel_report(
            connection,
            output_dir=output_dir,
            report_date=date(2026, 5, 3),
        )
    finally:
        connection.close()

    workbook = load_workbook(output_path)
    sheet = workbook["Trade Candidates"]
    validations = list(sheet.data_validations.dataValidation)
    validation_text = "\n".join(
        str(validation.formula1) for validation in validations
    )

    for decision in REVIEW_DECISION_VALUES:
        assert decision in validation_text

    assert validations


def test_trade_candidates_sheet_sorts_by_portfolio_status(
    tmp_path: Path,
) -> None:
    """Trade Candidates should put held names ahead of new names."""
    config_dir = tmp_path / "config"
    database_path = tmp_path / "data" / "market_sentinel.duckdb"
    output_dir = tmp_path / "outputs" / "excel"
    write_settings(config_dir, database_path)
    portfolio_dir = config_dir / "portfolio"
    portfolio_dir.mkdir(parents=True, exist_ok=True)
    portfolio_dir.joinpath("holdings.csv").write_text(
        "\n".join(
            [
                "ticker,name,market,quantity,average_cost,notes",
                "BBB,Example B,FTSE 350,25,100,Test holding",
            ]
        ),
        encoding="utf-8",
    )
    portfolio_dir.joinpath("watchlist.csv").write_text(
        "ticker,name,market,reason,notes\n",
        encoding="utf-8",
    )
    connection = open_duckdb_connection(config_dir)
    initialise_database_schema(connection)

    try:
        insert_report_data(connection)
        output_path = generate_excel_report(
            connection,
            output_dir=output_dir,
            report_date=date(2026, 5, 3),
            config_dir=config_dir,
        )
    finally:
        connection.close()

    workbook = load_workbook(output_path)
    sheet = workbook["Trade Candidates"]
    headers = [cell.value for cell in sheet[1]]
    portfolio_status_column = headers.index("Portfolio Status") + 1
    holding_quantity_column = headers.index("Holding Quantity") + 1
    summary_values = {
        workbook["Summary"][f"A{row}"].value: workbook["Summary"][f"B{row}"].value
        for row in range(1, workbook["Summary"].max_row + 1)
    }

    assert sheet["A2"].value == "BBB"
    assert sheet.cell(2, portfolio_status_column).value == "Held"
    assert sheet.cell(2, holding_quantity_column).value == "25"
    assert summary_values["Held candidates"] == 1
    assert summary_values["Watchlist candidates"] == 0
    assert summary_values["New candidates"] == 1
    assert summary_values["S&P 500 candidates"] == 1
    assert summary_values["FTSE 350 candidates"] == 1


def test_position_sizing_sheet_has_expected_labels_and_formulas(
    tmp_path: Path,
) -> None:
    """Position Sizing should provide simple editable inputs and formulas."""
    connection = open_test_database(tmp_path)
    output_dir = tmp_path / "outputs" / "excel"

    try:
        output_path = generate_excel_report(
            connection,
            output_dir=output_dir,
            report_date=date(2026, 5, 3),
        )
    finally:
        connection.close()

    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook["Position Sizing"]

    assert sheet["A2"].value == "Trading capital"
    assert sheet["A3"].value == "Risk per trade %"
    assert sheet["A4"].value == "Entry price"
    assert sheet["A5"].value == "Stop price"
    assert sheet["A6"].value == "Maximum £ risk"
    assert sheet["B6"].value == "=B2*B3"
    assert sheet["A7"].value == "Risk per unit/share/point"
    assert sheet["B7"].value == "=ABS(B4-B5)"
    assert sheet["A8"].value == "Suggested position size"
    assert sheet["B8"].value == '=IF(B7=0,"Check entry/stop",ROUNDDOWN(B6/B7,0))'
    assert "Position sizing is a planning calculation only" in sheet["B9"].value
    assert "fees, slippage, taxes, liquidity" in sheet["B9"].value
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref is not None


def test_bullish_position_sizing_calculation() -> None:
    """Bullish sizing should use entry minus 20-day low risk."""
    values = _position_sizing_values(
        {
            "signal_direction": "Bullish",
            "latest_close_price": 119.0,
        },
        {
            "20-day low": 100.0,
            "50-day SMA": 110.0,
        },
        {
            "trading_capital": 10000,
            "risk_per_trade_percent": 1,
            "default_stop_method": "20-day reference",
        },
    )

    assert values[:6] == (119.0, 100.0, 19.0, 100.0, 5, 595.0)
    assert "planning" in values[6].lower()


def test_bearish_position_sizing_calculation() -> None:
    """Bearish sizing should use 20-day high minus entry risk."""
    values = _position_sizing_values(
        {
            "signal_direction": "Bearish",
            "latest_close_price": 80.0,
        },
        {
            "20-day high": 95.0,
            "50-day SMA": 90.0,
        },
        {
            "trading_capital": 10000,
            "risk_per_trade_percent": 1,
            "default_stop_method": "20-day reference",
        },
    )

    assert values[:6] == (80.0, 95.0, 15.0, 100.0, 6, 480.0)


def test_invalid_position_sizing_stop_gives_safe_output() -> None:
    """Invalid stop levels should avoid producing a misleading size."""
    values = _position_sizing_values(
        {
            "signal_direction": "Bullish",
            "latest_close_price": 100.0,
        },
        {
            "20-day low": 105.0,
            "50-day SMA": 95.0,
        },
        {
            "trading_capital": 10000,
            "risk_per_trade_percent": 1,
            "default_stop_method": "20-day reference",
        },
    )

    assert values[2] == ""
    assert values[4] == ""
    assert values[5] == ""
    assert "Check stop" in values[6]


def test_apply_trade_candidate_position_sizing_uses_worksheet_columns() -> None:
    """Worksheet-level sizing should populate rows from actual headers."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Trade Candidates"
    sheet.append(
        [
            "Ticker",
            "Latest Close",
            "Direction",
            "20-day low reference",
            "20-day high reference",
            "50-day SMA reference",
            "Review decision",
            "Review notes",
            "Reviewed date",
        ]
    )
    sheet.append(["BULL", 120.0, "Bullish", 100.0, 130.0, 110.0, "", "", ""])
    sheet.append(["BEAR", 80.0, "Bearish", 70.0, 95.0, 90.0, "", "", ""])

    apply_trade_candidate_position_sizing(
        sheet,
        {
            "position_sizing_trading_capital": 10000,
            "position_sizing_risk_per_trade_percent": 1,
        },
    )
    headers = [cell.value for cell in sheet[1]]
    planned_entry_column = headers.index("Planned entry") + 1
    planned_stop_column = headers.index("Planned stop") + 1
    planned_risk_column = headers.index("Planned risk %") + 1
    risk_per_unit_column = headers.index("Risk per unit") + 1
    max_risk_column = headers.index("Max £ risk") + 1
    position_size_column = headers.index("Position size") + 1
    position_value_column = headers.index("Position value") + 1
    review_decision_column = headers.index("Review decision") + 1
    review_notes_column = headers.index("Review notes") + 1
    reviewed_date_column = headers.index("Reviewed date") + 1

    assert sheet.cell(2, planned_entry_column).value == '=IF(B2="","",B2)'
    assert sheet.cell(2, planned_stop_column).value == '=IF(D2="","",D2)'
    assert sheet.cell(3, planned_stop_column).value == '=IF(E3="","",E3)'
    assert sheet.cell(2, planned_risk_column).value == 0.01
    assert sheet.cell(2, risk_per_unit_column).value.startswith("=")
    assert sheet.cell(2, max_risk_column).value.startswith("=")
    assert sheet.cell(2, position_size_column).value.startswith("=")
    assert sheet.cell(2, position_value_column).value.startswith("=")
    assert sheet.cell(2, review_decision_column).value in (None, "")
    assert sheet.cell(2, review_notes_column).value in (None, "")
    assert sheet.cell(2, reviewed_date_column).value in (None, "")


def test_apply_trade_candidate_position_sizing_notes_missing_stop() -> None:
    """Rows with no stop source should get a Check stop note."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Trade Candidates"
    sheet.append(["Ticker", "Latest Close", "Direction"])
    sheet.append(["MISS", 120.0, "Bullish"])

    apply_trade_candidate_position_sizing(sheet, {})
    headers = [cell.value for cell in sheet[1]]
    planned_stop_column = headers.index("Planned stop") + 1
    position_size_column = headers.index("Position size") + 1
    note_column = headers.index("Position sizing note") + 1

    assert sheet.cell(2, planned_stop_column).value == ""
    assert sheet.cell(2, position_size_column).value.startswith("=")
    assert sheet.cell(2, note_column).value == "Check stop"


def test_trade_journal_sheet_has_expected_columns(tmp_path: Path) -> None:
    """Trade Journal should provide a beginner-friendly blank table."""
    connection = open_test_database(tmp_path)
    output_dir = tmp_path / "outputs" / "excel"

    try:
        output_path = generate_excel_report(
            connection,
            output_dir=output_dir,
            report_date=date(2026, 5, 3),
        )
    finally:
        connection.close()

    workbook = load_workbook(output_path)
    sheet = workbook["Trade Journal"]
    headers = [cell.value for cell in sheet[1]]

    assert headers == [
        "Date reviewed",
        "Ticker",
        "Company name",
        "Market",
        "Action grade",
        "Decision",
        "Entry planned",
        "Stop planned",
        "Risk %",
        "Trade taken?",
        "Entry date",
        "Exit date",
        "Exit reason",
        "Result",
        "Notes",
    ]
    assert sheet["F2"].value == "Watch | Paper trade | Trade | Ignore | Already held"
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref is not None


def test_generate_excel_report_creates_output_folder(tmp_path: Path) -> None:
    """Excel report generation should create the output folder automatically."""
    connection = open_test_database(tmp_path)
    output_dir = tmp_path / "new" / "outputs" / "excel"

    try:
        output_path = generate_excel_report(
            connection,
            output_dir=output_dir,
            report_date=date(2026, 5, 3),
        )
    finally:
        connection.close()

    assert output_dir.exists()
    assert output_path.exists()


def test_generate_excel_report_uses_configured_output_folder(
    tmp_path: Path,
    monkeypatch,
) -> None:
    """Excel reports should use configured output folders and expand ~."""
    monkeypatch.setenv("HOME", str(tmp_path))
    config_dir = tmp_path / "config"
    database_path = tmp_path / "data" / "market_sentinel.duckdb"
    write_report_settings(
        config_dir,
        database_path,
        "~/Library/CloudStorage/OneDrive-Personal/Finance/MarketSentinel/Excel",
    )
    connection = open_duckdb_connection(config_dir)
    initialise_database_schema(connection)

    try:
        output_path = generate_excel_report(
            connection,
            report_date=date(2026, 5, 3),
            config_dir=config_dir,
        )
    finally:
        connection.close()

    expected_dir = (
        tmp_path
        / "Library"
        / "CloudStorage"
        / "OneDrive-Personal"
        / "Finance"
        / "MarketSentinel"
        / "Excel"
    )

    assert output_path.parent == expected_dir
    assert output_path.exists()


def test_generate_excel_report_handles_missing_dividend_data(
    tmp_path: Path,
) -> None:
    """Dividend worksheets should still exist when there is no dividend data."""
    connection = open_test_database(tmp_path)
    output_dir = tmp_path / "outputs" / "excel"

    try:
        output_path = generate_excel_report(
            connection,
            output_dir=output_dir,
            report_date=date(2026, 5, 3),
        )
    finally:
        connection.close()

    workbook = load_workbook(output_path)

    assert workbook["Dividend Metrics"].max_row == 1
    assert workbook["High Dividend Stocks"].max_row == 1
    assert workbook["Dividend Risk Flags"].max_row == 1


def test_generate_excel_report_limits_large_moving_average_history(
    tmp_path: Path,
) -> None:
    """Large SMA history should be trimmed so Excel row limits are never hit."""
    config_dir = tmp_path / "config"
    database_path = tmp_path / "data" / "market_sentinel.duckdb"
    output_dir = tmp_path / "outputs" / "excel"
    write_report_settings(
        config_dir,
        database_path,
        str(output_dir),
        max_rows_per_sheet=5,
        recent_days=30,
    )
    connection = open_duckdb_connection(config_dir)
    initialise_database_schema(connection)

    try:
        connection.execute(
            """
            INSERT INTO securities (
                security_id,
                ticker,
                name,
                market,
                region,
                currency,
                sector
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            [
                1,
                "AAA",
                "Example A",
                "S&P 500",
                "United States",
                "USD",
                "Technology",
            ],
        )

        for index in range(20):
            connection.execute(
                """
                INSERT INTO moving_average_signals (
                    signal_id,
                    security_id,
                    signal_date,
                    moving_average_period_days,
                    moving_average_value,
                    signal_type
                )
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                [
                    index + 1,
                    1,
                    f"2026-05-{index + 1:02d}",
                    7,
                    100.0 + index,
                    "SMA",
                ],
            )

        output_path = generate_excel_report(
            connection,
            report_date=date(2026, 5, 3),
            config_dir=config_dir,
        )
    finally:
        connection.close()

    workbook = load_workbook(output_path)
    summary_text = "\n".join(
        str(workbook["Summary"][f"A{row}"].value)
        for row in range(1, workbook["Summary"].max_row + 1)
    )

    assert workbook["Moving Averages"].max_row == 2
    assert workbook["Moving Averages"]["C2"].value.date() == date(2026, 5, 20)
    assert workbook["Recent Moving Averages"].max_row == 6
    assert "Recent Moving Averages was limited to 5 rows" in summary_text


def test_generate_excel_report_script_uses_updated_report_code(
    tmp_path: Path,
    monkeypatch,
) -> None:
    """The script path should produce a workbook with all expected worksheets."""
    connection = open_test_database(tmp_path)
    output_dir = tmp_path / "outputs" / "excel"

    def fake_open_connection():
        return connection

    def fake_generate_report(connection_arg):
        return generate_excel_report(
            connection_arg,
            output_dir=output_dir,
            report_date=date(2026, 5, 3),
        )

    monkeypatch.setattr(
        generate_excel_report_script,
        "open_duckdb_connection",
        fake_open_connection,
    )
    monkeypatch.setattr(
        generate_excel_report_script,
        "generate_excel_report",
        fake_generate_report,
    )

    insert_report_data(connection)
    generate_excel_report_script.main()

    output_path = output_dir / "market_sentinel_report_2026-05-03.xlsx"
    workbook = load_workbook(output_path)

    assert workbook.sheetnames == EXPECTED_WORKSHEET_TITLES
