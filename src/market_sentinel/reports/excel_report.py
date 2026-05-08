"""Excel report generation for market-sentinel.

This module reads summary data from DuckDB and writes a beginner-friendly Excel
workbook using openpyxl. It does not create PDF reports.
"""

from datetime import date, datetime, timedelta
from math import floor
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import duckdb
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from market_sentinel.analytics.trade_candidates import (
    build_trade_candidate,
    portfolio_priority_rank,
)
from market_sentinel.analytics.crossovers import (
    DEFAULT_CROSSOVER_RECENT_DAYS,
    describe_crossover,
    format_days_since_crossover,
)
from market_sentinel.config.loader import load_named_config

DEFAULT_OUTPUT_DIR = Path("outputs") / "excel"
DEFAULT_MAX_ROWS_PER_SHEET = 50000
DEFAULT_MOVING_AVERAGE_RECENT_DAYS = 10
DEFAULT_POSITION_SIZING_TRADING_CAPITAL = 10000
DEFAULT_POSITION_SIZING_RISK_PER_TRADE_PERCENT = 1
DEFAULT_POSITION_SIZING_STOP_METHOD = "20-day reference"
EXCEL_MAX_DATA_ROWS = 1048575
EXPECTED_WORKSHEET_TITLES = [
    "Summary",
    "Securities",
    "Latest Prices",
    "Moving Averages",
    "Recent Moving Averages",
    "Recent Crossovers",
    "Crossover Signals",
    "Dividend Metrics",
    "High Dividend Stocks",
    "Dividend Risk Flags",
    "Trade Candidates",
    "Position Sizing",
    "Trade Journal",
]

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)
HEADER_FONT = Font(bold=True)
STRONG_BUY_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
STRONG_SELL_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
TRACK_ONLY_FILL = PatternFill(fill_type="solid", fgColor="F3F3F3")
REVIEW_DECISION_VALUES = [
    "Watch",
    "Paper trade",
    "Trade",
    "Ignore",
    "Already held",
]
TRADE_CANDIDATE_REVIEW_HEADERS = [
    "Review decision",
    "Review notes",
    "Planned entry",
    "Planned stop",
    "Planned risk %",
    "Position size",
    "Reviewed date",
]
TRADE_CANDIDATE_POSITION_HEADERS = [
    "Planning entry price",
    "Planning stop price",
    "Risk per unit",
    "Max £ risk",
    "Suggested position size",
    "Position value",
    "Position sizing note",
]


def default_report_filename(report_date: Optional[date] = None) -> str:
    """Return the default Excel report filename for a date."""
    selected_date = report_date or date.today()
    return f"market_sentinel_report_{selected_date.isoformat()}.xlsx"


def generate_excel_report(
    connection: duckdb.DuckDBPyConnection,
    output_dir: Optional[Path] = None,
    report_date: Optional[date] = None,
    config_dir: Optional[Path] = None,
) -> Path:
    """Generate an Excel report from the local DuckDB database."""
    selected_date = report_date or date.today()
    settings = _load_excel_settings(config_dir)
    target_dir = _resolve_excel_output_dir(output_dir, settings)
    output_path = target_dir / default_report_filename(selected_date)
    max_rows_per_sheet = _positive_int_setting(
        settings,
        "excel_max_rows_per_sheet",
        DEFAULT_MAX_ROWS_PER_SHEET,
    )
    max_rows_per_sheet = min(max_rows_per_sheet, EXCEL_MAX_DATA_ROWS)
    recent_days = _positive_int_setting(
        settings,
        "excel_moving_average_recent_days",
        DEFAULT_MOVING_AVERAGE_RECENT_DAYS,
    )
    crossover_recent_days = _positive_int_setting(
        settings,
        "crossover_recent_days",
        DEFAULT_CROSSOVER_RECENT_DAYS,
    )

    try:
        target_dir.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = EXPECTED_WORKSHEET_TITLES[0]
        limit_notes: List[str] = []

        report_sections = [
            ("Securities", _fetch_securities),
            ("Latest Prices", _fetch_latest_prices),
            ("Moving Averages", _fetch_moving_averages),
            (
                "Recent Moving Averages",
                lambda report_connection: _fetch_recent_moving_averages(
                    report_connection,
                    recent_days,
                ),
            ),
            (
                "Recent Crossovers",
                lambda report_connection: _fetch_crossover_signals(
                    report_connection,
                    selected_date,
                    crossover_recent_days,
                ),
            ),
            (
                "Crossover Signals",
                lambda report_connection: _fetch_crossover_signals(
                    report_connection,
                    selected_date,
                    None,
                ),
            ),
            ("Dividend Metrics", _fetch_dividend_metrics),
            ("High Dividend Stocks", _fetch_high_dividend_stocks),
            ("Dividend Risk Flags", _fetch_dividend_risk_flags),
        ]

        for sheet_title, fetch_data in report_sections:
            _write_table_sheet(
                workbook,
                sheet_title,
                fetch_data(connection),
                max_rows_per_sheet,
                limit_notes,
            )

        trade_candidate_data = _fetch_trade_candidates(
            connection,
            selected_date,
            crossover_recent_days,
            config_dir,
        )
        _write_trade_candidates_sheet(
            workbook,
            trade_candidate_data,
            max_rows_per_sheet,
            limit_notes,
        )
        _write_position_sizing_sheet(workbook, settings)
        _write_trade_journal_sheet(workbook)
        _write_summary_sheet(
            summary_sheet,
            connection,
            limit_notes,
            trade_candidate_data[1],
        )

        workbook.save(output_path)
    except duckdb.Error as error:
        raise RuntimeError(
            "Could not read report data from DuckDB. Check that the database "
            "is open and the required tables have been created. "
            f"Details: {error}"
        ) from error
    except OSError as error:
        raise RuntimeError(
            "Could not create or write to the Excel report folder. Check that "
            f"this path exists or can be created: {target_dir}."
        ) from error

    return output_path


def _resolve_excel_output_dir(
    output_dir: Optional[Path],
    settings: Dict[str, Any],
) -> Path:
    """Resolve the Excel output directory from arguments, settings, or fallback."""
    if output_dir is not None:
        return Path(output_dir).expanduser()

    configured_dir = settings.get("report_outputs", {}).get("excel_dir")

    if configured_dir:
        return Path(str(configured_dir)).expanduser()

    return DEFAULT_OUTPUT_DIR


def _load_excel_settings(config_dir: Optional[Path]) -> Dict[str, Any]:
    """Load settings for Excel reports, falling back to safe defaults."""
    try:
        loaded_settings = load_named_config("settings", config_dir)
    except FileNotFoundError:
        return {}

    return loaded_settings


def _positive_int_setting(
    settings: Dict[str, Any],
    setting_name: str,
    default_value: int,
) -> int:
    """Read a positive integer setting with a safe fallback."""
    raw_value = settings.get(setting_name, default_value)

    try:
        parsed_value = int(raw_value)
    except (TypeError, ValueError):
        return default_value

    if parsed_value < 1:
        return default_value

    return parsed_value


def _positive_float_setting(
    settings: Dict[str, Any],
    setting_name: str,
    default_value: float,
) -> float:
    """Read a positive float setting with a safe fallback."""
    raw_value = settings.get(setting_name, default_value)

    try:
        parsed_value = float(raw_value)
    except (TypeError, ValueError):
        return default_value

    if parsed_value <= 0:
        return default_value

    return parsed_value


def _position_sizing_settings(settings: Dict[str, Any]) -> Dict[str, Any]:
    """Return configurable default position sizing values."""
    return {
        "trading_capital": _positive_float_setting(
            settings,
            "position_sizing_trading_capital",
            DEFAULT_POSITION_SIZING_TRADING_CAPITAL,
        ),
        "risk_per_trade_percent": _positive_float_setting(
            settings,
            "position_sizing_risk_per_trade_percent",
            DEFAULT_POSITION_SIZING_RISK_PER_TRADE_PERCENT,
        ),
        "default_stop_method": str(
            settings.get(
                "position_sizing_default_stop_method",
                DEFAULT_POSITION_SIZING_STOP_METHOD,
            )
        ),
    }


def _write_summary_sheet(
    sheet,
    connection: duckdb.DuckDBPyConnection,
    limit_notes: Sequence[str],
    trade_candidate_rows: Optional[Sequence[Sequence[Any]]] = None,
) -> None:
    """Write the Summary worksheet."""
    rows = [
        ("Metric", "Value"),
        ("Securities", _count_rows(connection, "securities")),
        (
            "Securities With Dividend Metrics",
            _count_dividend_metric_securities(connection),
        ),
        ("Dividend Risk Flags", _count_dividend_risk_flags(connection)),
        ("Daily Price Rows", _count_rows(connection, "daily_prices")),
        ("Moving Average Rows", _count_rows(connection, "moving_average_signals")),
        ("Latest Price Date", _latest_price_date(connection) or "No prices yet"),
    ]

    if trade_candidate_rows is not None:
        rows.extend(_trade_candidate_summary_rows(trade_candidate_rows))

    if limit_notes:
        rows.append(("", ""))
        rows.append(("Report Notes", ""))
        for note in limit_notes:
            rows.append((note, ""))

    _write_rows(sheet, rows)


def _trade_candidate_summary_rows(
    trade_candidate_rows: Sequence[Sequence[Any]],
) -> List[Sequence[Any]]:
    """Return compact Trade Candidate counts for the Summary worksheet."""
    portfolio_counts = {
        "Held candidates": 0,
        "Watchlist candidates": 0,
        "New candidates": 0,
    }
    market_counts: Dict[str, int] = {}

    for row in trade_candidate_rows:
        market = str(row[2] or "Market unknown")
        portfolio_status = str(row[14] or "New")
        market_counts[market] = market_counts.get(market, 0) + 1

        if portfolio_status in {"Held", "Held + Watchlist"}:
            portfolio_counts["Held candidates"] += 1
        elif portfolio_status == "Watchlist":
            portfolio_counts["Watchlist candidates"] += 1
        else:
            portfolio_counts["New candidates"] += 1

    return [
        ("", ""),
        ("Trade Candidate Review", ""),
        ("Held candidates", portfolio_counts["Held candidates"]),
        ("Watchlist candidates", portfolio_counts["Watchlist candidates"]),
        ("New candidates", portfolio_counts["New candidates"]),
        ("S&P 500 candidates", market_counts.get("S&P 500", 0)),
        ("FTSE 350 candidates", market_counts.get("FTSE 350", 0)),
    ]


def _write_table_sheet(
    workbook: Workbook,
    title: str,
    table_data: Tuple[List[str], List[Sequence[Any]]],
    max_rows_per_sheet: int,
    limit_notes: List[str],
) -> None:
    """Create one worksheet from headers and rows."""
    sheet = workbook.create_sheet(title)
    headers, rows = table_data
    visible_rows = rows[:max_rows_per_sheet]

    if len(rows) > max_rows_per_sheet:
        limit_notes.append(
            f"{title} was limited to {max_rows_per_sheet} rows because the "
            "full dataset is larger than a readable Excel daily report."
        )

    _write_rows(sheet, [headers] + visible_rows)


def _write_trade_candidates_sheet(
    workbook: Workbook,
    table_data: Tuple[List[str], List[Sequence[Any]]],
    max_rows_per_sheet: int,
    limit_notes: List[str],
) -> None:
    """Create the Trade Candidates worksheet with review workflow columns."""
    sheet = workbook.create_sheet("Trade Candidates")
    headers, rows = table_data
    visible_rows = rows[:max_rows_per_sheet]

    if len(rows) > max_rows_per_sheet:
        limit_notes.append(
            f"Trade Candidates was limited to {max_rows_per_sheet} rows because "
            "the full dataset is larger than a readable Excel daily report."
        )

    workflow_headers = headers + TRADE_CANDIDATE_REVIEW_HEADERS
    workflow_rows = [
        list(row) + [""] * len(TRADE_CANDIDATE_REVIEW_HEADERS)
        for row in visible_rows
    ]
    _write_rows(sheet, [workflow_headers] + workflow_rows)
    _add_review_decision_validation(sheet, workflow_headers)
    _add_trade_candidate_conditional_formatting(sheet, workflow_headers)
    _format_trade_candidate_position_columns(sheet, workflow_headers)
    _auto_size_columns(sheet)


def _write_rows(sheet, rows: Iterable[Sequence[Any]]) -> None:
    """Write rows and apply simple formatting."""
    for row in rows:
        sheet.append(list(row))

    if sheet.max_row >= 1:
        for cell in sheet[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    _auto_size_columns(sheet)


def _add_review_decision_validation(sheet, headers: Sequence[str]) -> None:
    """Add a simple dropdown for daily candidate review decisions."""
    if "Review decision" not in headers:
        return

    column_index = headers.index("Review decision") + 1
    column_letter = get_column_letter(column_index)
    formula_values = ",".join(REVIEW_DECISION_VALUES)
    validation = DataValidation(
        type="list",
        formula1=f'"{formula_values}"',
        allow_blank=True,
    )
    validation.error = "Choose one of the suggested review decisions."
    validation.errorTitle = "Review decision"
    validation.prompt = "Use Trade Candidates for daily review."
    validation.promptTitle = "Review decision"
    sheet.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}{max(sheet.max_row, 1000)}")


def _add_trade_candidate_conditional_formatting(
    sheet,
    headers: Sequence[str],
) -> None:
    """Highlight setup grades without changing candidate scoring."""
    if "Action Grade" not in headers or sheet.max_row < 2:
        return

    grade_column = get_column_letter(headers.index("Action Grade") + 1)
    last_column = get_column_letter(len(headers))
    data_range = f"A2:{last_column}{sheet.max_row}"
    grade_reference = f"${grade_column}2"
    sheet.conditional_formatting.add(
        data_range,
        FormulaRule(
            formula=[f'{grade_reference}="Strong Buy Setup"'],
            fill=STRONG_BUY_FILL,
        ),
    )
    sheet.conditional_formatting.add(
        data_range,
        FormulaRule(
            formula=[f'{grade_reference}="Strong Sell Setup"'],
            fill=STRONG_SELL_FILL,
        ),
    )
    sheet.conditional_formatting.add(
        data_range,
        FormulaRule(
            formula=[f'{grade_reference}="Track Only"'],
            fill=TRACK_ONLY_FILL,
        ),
    )


def _format_trade_candidate_position_columns(sheet, headers: Sequence[str]) -> None:
    """Apply readable number formats to position planning columns."""
    currency_headers = {
        "Planning entry price",
        "Planning stop price",
        "Risk per unit",
        "Max £ risk",
        "Position value",
    }
    integer_headers = {"Suggested position size"}

    for column_index, header in enumerate(headers, start=1):
        if header in currency_headers:
            number_format = "£#,##0.00"
        elif header in integer_headers:
            number_format = "0"
        else:
            continue

        for row in range(2, sheet.max_row + 1):
            sheet.cell(row, column_index).number_format = number_format


def _write_position_sizing_sheet(
    workbook: Workbook,
    settings: Optional[Dict[str, Any]] = None,
) -> None:
    """Create a beginner-friendly position sizing calculator sheet."""
    sizing_settings = _position_sizing_settings(settings or {})
    trading_capital = sizing_settings["trading_capital"]
    risk_decimal = sizing_settings["risk_per_trade_percent"] / 100
    sheet = workbook.create_sheet("Position Sizing")
    rows = [
        ("Planning calculator", "Value", "Notes"),
        (
            "Trading capital",
            trading_capital,
            "Editable example default used for Trade Candidates planning",
        ),
        (
            "Risk per trade %",
            risk_decimal,
            "Editable example default used for Trade Candidates planning",
        ),
        ("Entry price", 100, "Example only - edit this input"),
        ("Stop price", 95, "Example only - edit this input"),
        ("Maximum £ risk", "=B2*B3", "Trading capital times risk per trade"),
        ("Risk per unit/share/point", "=ABS(B4-B5)", "Entry price minus stop price"),
        (
            "Suggested position size",
            '=IF(B7=0,"Check entry/stop",ROUNDDOWN(B6/B7,0))',
            "Planning size only",
        ),
        (
            "Note",
            (
                "Position sizing is a planning calculation only. It does not "
                "account for fees, slippage, taxes, liquidity, or personal "
                "circumstances."
            ),
            "",
        ),
    ]
    _write_rows(sheet, rows)
    sheet["B3"].number_format = "0.00%"
    sheet["B6"].number_format = "£#,##0.00"
    sheet["B7"].number_format = "0.00"


def _write_trade_journal_sheet(workbook: Workbook) -> None:
    """Create a blank trade journal worksheet."""
    sheet = workbook.create_sheet("Trade Journal")
    headers = [
        "Date reviewed",
        "Ticker",
        "Company name",
        "Market",
        "Action grade",
        "Decision",
        "Entry planned",
        "Stop planned",
        "Risk %",
        "Trade taken?",
        "Entry date",
        "Exit date",
        "Exit reason",
        "Result",
        "Notes",
    ]
    rows = [
        headers,
        [
            "Suggested Decision values:",
            "",
            "",
            "",
            "",
            "Watch | Paper trade | Trade | Ignore | Already held",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
    ]
    _write_rows(sheet, rows)


def _auto_size_columns(sheet) -> None:
    """Set readable column widths."""
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        sheet.column_dimensions[column_letter].width = min(max_length + 2, 40)


def _fetch_securities(
    connection: duckdb.DuckDBPyConnection,
) -> Tuple[List[str], List[Sequence[Any]]]:
    """Fetch securities for the report."""
    headers = ["Ticker", "Name", "Market", "Region", "Currency", "Sector"]
    rows = connection.execute(
        f"""
        SELECT
            ticker,
            name,
            {_market_expression("securities")} AS market,
            region,
            currency,
            sector
        FROM securities
        ORDER BY ticker
        """
    ).fetchall()
    return headers, rows


def _market_expression(alias: str = "securities") -> str:
    """Return SQL that displays a readable market with a ticker fallback."""
    return (
        "CASE "
        f"WHEN {alias}.market IS NOT NULL AND TRIM({alias}.market) <> '' "
        f"THEN {alias}.market "
        f"WHEN UPPER({alias}.ticker) LIKE '%.L' THEN 'FTSE 350' "
        "ELSE 'S&P 500' "
        "END"
    )


def _fetch_latest_prices(
    connection: duckdb.DuckDBPyConnection,
) -> Tuple[List[str], List[Sequence[Any]]]:
    """Fetch each security's latest daily price."""
    headers = [
        "Ticker",
        "Market",
        "Price Date",
        "Open",
        "High",
        "Low",
        "Close",
        "Adjusted Close",
        "Volume",
    ]
    rows = connection.execute(
        f"""
        SELECT
            securities.ticker,
            {_market_expression("securities")} AS market,
            latest_prices.price_date,
            latest_prices.open_price,
            latest_prices.high_price,
            latest_prices.low_price,
            latest_prices.close_price,
            latest_prices.adjusted_close_price,
            latest_prices.volume
        FROM securities
        INNER JOIN daily_prices AS latest_prices
            ON securities.security_id = latest_prices.security_id
        INNER JOIN (
            SELECT security_id, MAX(price_date) AS latest_date
            FROM daily_prices
            GROUP BY security_id
        ) AS latest_dates
            ON latest_prices.security_id = latest_dates.security_id
           AND latest_prices.price_date = latest_dates.latest_date
        ORDER BY securities.ticker
        """
    ).fetchall()
    return headers, rows


def _fetch_moving_averages(
    connection: duckdb.DuckDBPyConnection,
) -> Tuple[List[str], List[Sequence[Any]]]:
    """Fetch the latest SMA value for each ticker and moving-average period."""
    headers = ["Ticker", "Market", "Signal Date", "Period Days", "SMA Value"]
    rows = connection.execute(
        f"""
        WITH latest_signal_dates AS (
            SELECT
                security_id,
                moving_average_period_days,
                MAX(signal_date) AS latest_signal_date
            FROM moving_average_signals
            WHERE signal_type = 'SMA'
            GROUP BY security_id, moving_average_period_days
        )
        SELECT
            securities.ticker,
            {_market_expression("securities")} AS market,
            signals.signal_date,
            signals.moving_average_period_days,
            signals.moving_average_value
        FROM moving_average_signals AS signals
        INNER JOIN latest_signal_dates
            ON signals.security_id = latest_signal_dates.security_id
           AND signals.moving_average_period_days =
               latest_signal_dates.moving_average_period_days
           AND signals.signal_date = latest_signal_dates.latest_signal_date
        INNER JOIN securities
            ON signals.security_id = securities.security_id
        WHERE signals.signal_type = 'SMA'
        ORDER BY securities.ticker, signals.moving_average_period_days
        """
    ).fetchall()
    return headers, rows


def _fetch_recent_moving_averages(
    connection: duckdb.DuckDBPyConnection,
    recent_days: int,
) -> Tuple[List[str], List[Sequence[Any]]]:
    """Fetch recent SMA history for quick checking without exporting all history."""
    headers = ["Ticker", "Market", "Signal Date", "Period Days", "SMA Value"]
    latest_signal_date = _latest_signal_date(connection, "SMA")

    if latest_signal_date is None:
        return headers, []

    cutoff_date = latest_signal_date - timedelta(days=recent_days - 1)
    rows = connection.execute(
        f"""
        SELECT
            securities.ticker,
            {_market_expression("securities")} AS market,
            signals.signal_date,
            signals.moving_average_period_days,
            signals.moving_average_value
        FROM moving_average_signals AS signals
        INNER JOIN securities
            ON signals.security_id = securities.security_id
        WHERE signals.signal_type = 'SMA'
          AND signals.signal_date >= ?
        ORDER BY signals.signal_date DESC,
                 securities.ticker,
                 signals.moving_average_period_days
        """,
        [cutoff_date],
    ).fetchall()
    return headers, rows


def _fetch_crossover_signals(
    connection: duckdb.DuckDBPyConnection,
    report_date: date,
    recent_days: Optional[int],
) -> Tuple[List[str], List[Sequence[Any]]]:
    """Fetch crossover signal rows, optionally filtered to a recent window."""
    headers = [
        "Ticker",
        "Company Name",
        "Market",
        "Signal Direction",
        "Signal Description",
        "Crossover Date",
        "Days Since Crossover",
    ]
    filters = ["signals.signal_type IN ('BULLISH_CROSSOVER', 'BEARISH_CROSSOVER')"]
    parameters: List[Any] = []

    if recent_days is not None:
        filters.append("signals.signal_date >= ?")
        filters.append("signals.signal_date <= ?")
        parameters.extend(
            [
                report_date - timedelta(days=recent_days),
                report_date,
            ]
        )

    where_clause = " AND ".join(filters)
    rows = connection.execute(
        f"""
        SELECT
            securities.ticker,
            securities.name,
            {_market_expression("securities")} AS market,
            signals.signal_date,
            signals.moving_average_period_days,
            signals.comparison_period_days,
            signals.crossover_direction
        FROM moving_average_signals AS signals
        INNER JOIN securities
            ON signals.security_id = securities.security_id
        WHERE {where_clause}
        ORDER BY
            CASE
                WHEN signals.signal_type = 'BULLISH_CROSSOVER' THEN 0
                ELSE 1
            END,
            signals.signal_date DESC,
            securities.ticker
        """,
        parameters,
    ).fetchall()
    formatted_rows = [
        (
            row[0],
            row[1],
            row[2],
            _friendly_direction(row[6]),
            describe_crossover(row[4], row[5], row[6]),
            row[3],
            format_days_since_crossover(row[3], report_date),
        )
        for row in rows
    ]
    return headers, formatted_rows


def _fetch_dividend_metrics(
    connection: duckdb.DuckDBPyConnection,
) -> Tuple[List[str], List[Sequence[Any]]]:
    """Fetch dividend metrics rows."""
    headers = [
        "Ticker",
        "Market",
        "Metric Date",
        "Trailing 12M Dividend",
        "Dividend Yield",
        "Annual Cash Per 10000",
        "Risk Flag",
        "Risk Reason",
    ]
    rows = connection.execute(
        f"""
        SELECT
            securities.ticker,
            {_market_expression("securities")} AS market,
            metrics.metric_date,
            metrics.trailing_annual_dividend,
            metrics.dividend_yield,
            metrics.annual_dividend_cash_per_10000,
            metrics.dividend_risk_flag,
            metrics.dividend_risk_reason
        FROM dividend_metrics AS metrics
        INNER JOIN securities
            ON metrics.security_id = securities.security_id
        ORDER BY securities.ticker, metrics.metric_date DESC
        """
    ).fetchall()
    return headers, rows


def _friendly_direction(crossover_direction: Any) -> str:
    """Return a readable crossover direction label."""
    if crossover_direction == "BULLISH_CROSSOVER":
        return "Bullish"

    if crossover_direction == "BEARISH_CROSSOVER":
        return "Bearish"

    return str(crossover_direction or "")


def _fetch_high_dividend_stocks(
    connection: duckdb.DuckDBPyConnection,
) -> Tuple[List[str], List[Sequence[Any]]]:
    """Fetch dividend metrics sorted by yield descending."""
    headers = [
        "Ticker",
        "Market",
        "Metric Date",
        "Dividend Yield",
        "Trailing 12M Dividend",
        "Annual Cash Per 10000",
        "Risk Flag",
        "Risk Reason",
    ]
    rows = connection.execute(
        f"""
        SELECT
            securities.ticker,
            {_market_expression("securities")} AS market,
            metrics.metric_date,
            metrics.dividend_yield,
            metrics.trailing_annual_dividend,
            metrics.annual_dividend_cash_per_10000,
            metrics.dividend_risk_flag,
            metrics.dividend_risk_reason
        FROM dividend_metrics AS metrics
        INNER JOIN securities
            ON metrics.security_id = securities.security_id
        WHERE metrics.dividend_yield IS NOT NULL
        ORDER BY metrics.dividend_yield DESC, securities.ticker
        """
    ).fetchall()
    return headers, rows


def _fetch_dividend_risk_flags(
    connection: duckdb.DuckDBPyConnection,
) -> Tuple[List[str], List[Sequence[Any]]]:
    """Fetch dividend risk flag rows."""
    headers = [
        "Ticker",
        "Market",
        "Metric Date",
        "Dividend Yield",
        "Risk Flag",
        "Risk Reason",
    ]
    rows = connection.execute(
        f"""
        SELECT
            securities.ticker,
            {_market_expression("securities")} AS market,
            metrics.metric_date,
            metrics.dividend_yield,
            metrics.dividend_risk_flag,
            metrics.dividend_risk_reason
        FROM dividend_metrics AS metrics
        INNER JOIN securities
            ON metrics.security_id = securities.security_id
        WHERE metrics.dividend_risk_flag IS NOT NULL
        ORDER BY metrics.dividend_yield DESC, securities.ticker
        """
    ).fetchall()
    return headers, rows


def _fetch_trade_candidates(
    connection: duckdb.DuckDBPyConnection,
    report_date: date,
    recent_days: int,
    config_dir: Optional[Path],
) -> Tuple[List[str], List[Sequence[Any]]]:
    """Fetch all recent crossover candidates for decision review."""
    headers = [
        "Ticker",
        "Company Name",
        "Market",
        "Direction",
        "Action Grade",
        "Score",
        "Crossover Date",
        "Days Since Crossover",
        "Signal Description",
        "Latest Close",
        "50-day SMA Reference",
        "20-day Low/High Reference",
        "20% Trailing Reference",
        "Dividend Risk Flag",
        "Portfolio Status",
        "Holding Quantity",
        "Watchlist Reason",
        "Risk Notes",
        *TRADE_CANDIDATE_POSITION_HEADERS,
    ]
    sizing_settings = _position_sizing_settings(_load_excel_settings(config_dir))
    raw_rows = connection.execute(
        """
        SELECT
            securities.ticker,
            signals.signal_date,
            signals.moving_average_period_days,
            signals.comparison_period_days,
            signals.signal_type
        FROM moving_average_signals AS signals
        INNER JOIN securities
            ON signals.security_id = securities.security_id
        WHERE signals.signal_type IN ('BULLISH_CROSSOVER', 'BEARISH_CROSSOVER')
          AND signals.signal_date >= ?
          AND signals.signal_date <= ?
        ORDER BY signals.signal_date DESC, securities.ticker
        """,
        [report_date - timedelta(days=recent_days), report_date],
    ).fetchall()
    rows = []

    for row in raw_rows:
        signal = {
            "direction": _friendly_direction(row[4]),
            "trend_description": describe_crossover(row[2], row[3], row[4]),
            "crossover_date": row[1],
            "days_since_crossover": format_days_since_crossover(row[1], report_date),
        }
        candidate = build_trade_candidate(
            connection,
            row[0],
            signal,
            config_dir=config_dir,
        )

        if candidate is None:
            continue

        review_levels = candidate.get("review_levels", {})
        position_sizing = _position_sizing_values(
            candidate,
            review_levels,
            sizing_settings,
        )
        rows.append(
            (
                candidate.get("ticker"),
                candidate.get("company_name"),
                candidate.get("market"),
                candidate.get("signal_direction"),
                candidate.get("action_grade"),
                candidate.get("score"),
                candidate.get("crossover_date"),
                candidate.get("days_since_crossover"),
                candidate.get("signal_description"),
                candidate.get("latest_close_price"),
                review_levels.get("50-day SMA"),
                _twenty_day_reference(candidate),
                review_levels.get("20% trailing reference"),
                candidate.get("dividend_risk_flag") or "",
                candidate.get("portfolio_status") or "New",
                candidate.get("holding_quantity") or "",
                candidate.get("watchlist_reason") or "",
                " | ".join(candidate.get("risk_notes", [])),
                *position_sizing,
            )
        )

    rows.sort(key=_trade_candidate_sort_key)
    return headers, rows


def _position_sizing_values(
    candidate: Dict[str, Any],
    review_levels: Dict[str, Any],
    sizing_settings: Dict[str, Any],
) -> Tuple[Any, Any, Any, Any, Any, Any, str]:
    """Calculate risk-planning position sizing values for one candidate."""
    entry_price = candidate.get("latest_close_price")
    direction = candidate.get("signal_direction")
    stop_price, stop_label = _planning_stop_price(direction, review_levels)
    max_risk = (
        sizing_settings["trading_capital"]
        * sizing_settings["risk_per_trade_percent"]
        / 100
    )
    risk_per_unit = _risk_per_unit(direction, entry_price, stop_price)

    if risk_per_unit is None or risk_per_unit <= 0:
        return (
            entry_price,
            stop_price,
            "",
            round(max_risk, 2),
            "",
            "",
            "Check stop - risk per unit is not positive.",
        )

    suggested_size = floor(max_risk / risk_per_unit)
    position_value = suggested_size * float(entry_price)
    note = (
        f"Risk planning using {stop_label}. Planning calculation only; "
        "does not account for fees, slippage, taxes, liquidity, or personal "
        "circumstances."
    )

    return (
        entry_price,
        stop_price,
        round(risk_per_unit, 2),
        round(max_risk, 2),
        suggested_size,
        round(position_value, 2),
        note,
    )


def _planning_stop_price(
    direction: Any,
    review_levels: Dict[str, Any],
) -> Tuple[Any, str]:
    """Return the default planning stop price and label for a direction."""
    if direction == "Bullish":
        if review_levels.get("20-day low") is not None:
            return review_levels.get("20-day low"), "20-day reference"
        if review_levels.get("50-day SMA") is not None:
            return review_levels.get("50-day SMA"), "50-day SMA fallback"

    if direction == "Bearish":
        if review_levels.get("20-day high") is not None:
            return review_levels.get("20-day high"), "20-day reference"
        if review_levels.get("50-day SMA") is not None:
            return review_levels.get("50-day SMA"), "50-day SMA fallback"

    return None, "no available stop reference"


def _risk_per_unit(
    direction: Any,
    entry_price: Any,
    stop_price: Any,
) -> Optional[float]:
    """Calculate directional risk per unit when prices are valid."""
    try:
        entry = float(entry_price)
        stop = float(stop_price)
    except (TypeError, ValueError):
        return None

    if direction == "Bullish":
        return entry - stop

    if direction == "Bearish":
        return stop - entry

    return None


def _twenty_day_reference(candidate: Dict[str, Any]) -> Any:
    """Return the candidate's directional 20-day high or low reference."""
    review_levels = candidate.get("review_levels", {})

    if candidate.get("signal_direction") == "Bullish":
        return review_levels.get("20-day low")

    if candidate.get("signal_direction") == "Bearish":
        return review_levels.get("20-day high")

    return review_levels.get("20-day extreme")


def _trade_candidate_sort_key(row: Sequence[Any]) -> tuple:
    """Sort Trade Candidates rows by portfolio-aware review priority."""
    grade = row[4]
    score = row[5] or 0
    market = row[2] or ""
    portfolio_status = row[14] or "New"

    return (
        portfolio_priority_rank(portfolio_status),
        -score,
        _trade_candidate_grade_rank(grade),
        market,
        row[0] or "",
    )


def _trade_candidate_grade_rank(grade: Any) -> int:
    """Return workbook sort rank for action grades."""
    ranks = {
        "Strong Buy Setup": 0,
        "Buy Setup": 1,
        "Track Only": 2,
        "Sell Setup": 3,
        "Strong Sell Setup": 4,
    }
    return ranks.get(str(grade), 2)


def _count_rows(connection: duckdb.DuckDBPyConnection, table_name: str) -> int:
    """Count rows in a known project table."""
    allowed_tables = {
        "securities",
        "daily_prices",
        "moving_average_signals",
        "dividend_metrics",
    }

    if table_name not in allowed_tables:
        raise ValueError(f"Unknown report table: {table_name}")

    return connection.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]


def _count_dividend_metric_securities(connection: duckdb.DuckDBPyConnection) -> int:
    """Count securities that have at least one dividend metric row."""
    return connection.execute(
        """
        SELECT COUNT(DISTINCT security_id)
        FROM dividend_metrics
        """
    ).fetchone()[0]


def _count_dividend_risk_flags(connection: duckdb.DuckDBPyConnection) -> int:
    """Count dividend metric rows with risk flags."""
    return connection.execute(
        """
        SELECT COUNT(*)
        FROM dividend_metrics
        WHERE dividend_risk_flag IS NOT NULL
        """
    ).fetchone()[0]


def _latest_price_date(connection: duckdb.DuckDBPyConnection) -> Any:
    """Return the latest daily price date."""
    return connection.execute("SELECT MAX(price_date) FROM daily_prices").fetchone()[0]


def _latest_signal_date(
    connection: duckdb.DuckDBPyConnection,
    signal_type: Any,
) -> Optional[date]:
    """Return the latest moving-average signal date for one or more signal types."""
    if isinstance(signal_type, tuple):
        placeholders = ", ".join(["?"] * len(signal_type))
        query = (
            "SELECT MAX(signal_date) FROM moving_average_signals "
            f"WHERE signal_type IN ({placeholders})"
        )
        latest_value = connection.execute(query, list(signal_type)).fetchone()[0]
    else:
        latest_value = connection.execute(
            """
            SELECT MAX(signal_date)
            FROM moving_average_signals
            WHERE signal_type = ?
            """,
            [signal_type],
        ).fetchone()[0]

    if latest_value is None:
        return None

    if isinstance(latest_value, datetime):
        return latest_value.date()

    if isinstance(latest_value, date):
        return latest_value

    return date.fromisoformat(str(latest_value)[:10])
